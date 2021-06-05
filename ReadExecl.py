import openpyxl
path = "C:\\Users\\addyf\Downloads\MnS.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active  #workbook.get_sheet_by_name("MnS.xlsx")

rows = sheet.max_row
cols = sheet.max_column
print(rows) #23
print(cols) #19
for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(row=r, column=c).value, end="                        ")
    print()






